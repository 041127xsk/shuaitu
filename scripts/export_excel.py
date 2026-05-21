"""
export_excel.py - 导出战报数据到 Excel（自动列宽）
=================================================
从 stzbHelper 数据库导出战报数据到 Excel 文件，自动调整列宽。

用法：
    python scripts/export_excel.py                          # 默认导出到 data/battle_reports.xlsx
    python scripts/export_excel.py -o data/my_report.xlsx   # 指定输出路径
    python scripts/export_excel.py --filter-valid            # 只导出有效战报（攻方3武将）
    python scripts/export_excel.py --filter-no-npc           # 排除 NPC 战斗
"""
from __future__ import annotations

import argparse
import sqlite3
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]

# 战报表字段映射（字段名 → 中文表头）
BATTLE_REPORT_FIELDS = {
    "battle_id": "战斗ID",
    "battle_time": "战斗时间",
    "wid_name": "战斗地点",
    "attack_name": "进攻方",
    "attack_union_name": "进攻方同盟",
    "defend_name": "防守方",
    "defend_union_name": "防守方同盟",
    "attack_hero1_id": "攻方大营ID",
    "attack_hero2_id": "攻方中军ID",
    "attack_hero3_id": "攻方前锋ID",
    "attack_hero1_level": "攻方大营等级",
    "attack_hero2_level": "攻方中军等级",
    "attack_hero3_level": "攻方前锋等级",
    "attack_hero1_star": "攻方大营红度",
    "attack_hero2_star": "攻方中军红度",
    "attack_hero3_star": "攻方前锋红度",
    "attack_total_star": "攻方总红度",
    "defend_hero1_id": "守方大营ID",
    "defend_hero2_id": "守方中军ID",
    "defend_hero3_id": "守方前锋ID",
    "defend_hero1_level": "守方大营等级",
    "defend_hero2_level": "守方中军等级",
    "defend_hero3_level": "守方前锋等级",
    "defend_hero1_star": "守方大营红度",
    "defend_hero2_star": "守方中军红度",
    "defend_hero3_star": "守方前锋红度",
    "defend_total_star": "守方总红度",
    "attack_hp": "攻方兵力",
    "defend_hp": "守方兵力",
    "npc": "NPC",
    "result": "结果",
}

# 同盟成员字段映射
TEAM_MEMBER_FIELDS = {
    "member_id": "成员ID",
    "name": "名称",
    "contribute_total": "总贡献",
    "contribute_week": "周贡献",
    "power": "势力值",
    "wu": "武勋",
    "group_name": "分组",
    "join_time": "加入时间",
}


def auto_fit_column_width(ws):
    """自动调整列宽（基于内容最大长度）"""
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in col_cells:
            if cell.value:
                # 中文字符算 2 个宽度
                value_str = str(cell.value)
                length = 0
                for char in value_str:
                    if '\u4e00' <= char <= '\u9fff' or '\u3000' <= char <= '\u303f':
                        length += 2
                    else:
                        length += 1
                max_length = max(max_length, length)
        
        # 设置列宽（最小 8，最大 40）
        adjusted_width = min(max(max_length + 2, 8), 40)
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = adjusted_width


def apply_header_style(ws):
    """应用表头样式"""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border


def apply_data_style(ws):
    """应用数据区域样式"""
    data_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = data_alignment
            cell.border = thin_border


def export_battle_reports(conn: sqlite3.Connection, filter_valid: bool = False, filter_no_npc: bool = False) -> openpyxl.Workbook:
    """导出战报数据到 Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "战报数据"
    
    # 构建查询
    fields = list(BATTLE_REPORT_FIELDS.keys())
    field_names = list(BATTLE_REPORT_FIELDS.values())
    
    where_clauses = []
    if filter_valid:
        where_clauses.append("attack_hero1_id != 0 AND attack_hero2_id != 0 AND attack_hero3_id != 0")
    if filter_no_npc:
        where_clauses.append("npc != 1")
    
    where_sql = " WHERE " + " AND ".join(where_clauses) if where_clauses else ""
    
    cursor = conn.cursor()
    cursor.execute(f"SELECT {', '.join(fields)} FROM stzb_battle_reports{where_sql} ORDER BY battle_time DESC")
    rows = cursor.fetchall()
    
    # 写入表头
    ws.append(field_names)
    
    # 写入数据
    for row in rows:
        ws.append(list(row))
    
    # 应用样式
    apply_header_style(ws)
    apply_data_style(ws)
    auto_fit_column_width(ws)
    
    # 冻结首行
    ws.freeze_panes = "A2"
    
    return wb


def export_team_members(conn: sqlite3.Connection) -> openpyxl.Workbook:
    """导出同盟成员数据到 Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "同盟成员"
    
    fields = list(TEAM_MEMBER_FIELDS.keys())
    field_names = list(TEAM_MEMBER_FIELDS.values())
    
    cursor = conn.cursor()
    cursor.execute(f"SELECT {', '.join(fields)} FROM stzb_team_members ORDER BY power DESC")
    rows = cursor.fetchall()
    
    # 写入表头
    ws.append(field_names)
    
    # 写入数据
    for row in rows:
        ws.append(list(row))
    
    # 应用样式
    apply_header_style(ws)
    apply_data_style(ws)
    auto_fit_column_width(ws)
    
    # 冻结首行
    ws.freeze_panes = "A2"
    
    return wb


def main():
    parser = argparse.ArgumentParser(description="导出战报数据到 Excel")
    parser.add_argument("-o", "--output", help="输出文件路径（默认 data/battle_reports.xlsx）")
    parser.add_argument("--filter-valid", action="store_true", help="只导出有效战报（攻方3武将）")
    parser.add_argument("--filter-no-npc", action="store_true", help="排除 NPC 战斗")
    parser.add_argument("--all", action="store_true", help="导出所有表（战报+成员）")
    args = parser.parse_args()
    
    # 查找数据库
    db_files = list(ROOT.glob("data/heroes.db"))
    if not db_files:
        print("[ERROR] 数据库不存在，请先运行 import_stzb_helper.py")
        return 1
    
    db_path = db_files[0]
    conn = sqlite3.connect(str(db_path))
    
    # 默认输出路径
    output_path = args.output or str(ROOT / "data" / "battle_reports.xlsx")
    
    print(f"数据库: {db_path}")
    print(f"输出: {output_path}")
    
    if args.all:
        # 导出所有表到不同 sheet
        wb = openpyxl.Workbook()
        
        # 战报表
        ws1 = wb.active
        ws1.title = "战报数据"
        wb_reports = export_battle_reports(conn, args.filter_valid, args.filter_no_npc)
        for row in wb_reports.active.iter_rows():
            ws1.append([cell.value for cell in row])
        apply_header_style(ws1)
        apply_data_style(ws1)
        auto_fit_column_width(ws1)
        ws1.freeze_panes = "A2"
        
        # 成员表
        ws2 = wb.create_sheet("同盟成员")
        wb_members = export_team_members(conn)
        for row in wb_members.active.iter_rows():
            ws2.append([cell.value for cell in row])
        apply_header_style(ws2)
        apply_data_style(ws2)
        auto_fit_column_width(ws2)
        ws2.freeze_panes = "A2"
        
        wb.save(output_path)
        print(f"已导出战报+成员到 {output_path}")
    else:
        # 只导出战报
        wb = export_battle_reports(conn, args.filter_valid, args.filter_no_npc)
        wb.save(output_path)
        print(f"已导出战报到 {output_path}")
    
    conn.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
