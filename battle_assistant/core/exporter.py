"""
战报助手 - Excel 导出模块
"""
import sqlite3
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# 战报表字段映射
BATTLE_REPORT_FIELDS = {
    "battle_id": "战斗ID",
    "battle_time": "战斗时间",
    "wid_name": "战斗地点",
    "attack_name": "进攻方",
    "attack_union_name": "进攻方同盟",
    "defend_name": "防守方",
    "defend_union_name": "防守方同盟",
    "attack_hero1_id": "攻方大营ID",
    "attack_hero2_id": "攻方中军ID",
    "attack_hero3_id": "攻方前锋ID",
    "attack_hero1_level": "攻方大营等级",
    "attack_hero2_level": "攻方中军等级",
    "attack_hero3_level": "攻方前锋等级",
    "attack_hero1_star": "攻方大营红度",
    "attack_hero2_star": "攻方中军红度",
    "attack_hero3_star": "攻方前锋红度",
    "attack_total_star": "攻方总红度",
    "defend_hero1_id": "守方大营ID",
    "defend_hero2_id": "守方中军ID",
    "defend_hero3_id": "守方前锋ID",
    "defend_hero1_level": "守方大营等级",
    "defend_hero2_level": "守方中军等级",
    "defend_hero3_level": "守方前锋等级",
    "defend_hero1_star": "守方大营红度",
    "defend_hero2_star": "守方中军红度",
    "defend_hero3_star": "守方前锋红度",
    "defend_total_star": "守方总红度",
    "attack_hp": "攻方兵力",
    "defend_hp": "守方兵力",
    "npc": "NPC",
    "result": "结果",
}

TEAM_MEMBER_FIELDS = {
    "member_id": "成员ID",
    "name": "名称",
    "contribute_total": "总贡献",
    "contribute_week": "周贡献",
    "power": "势力值",
    "wu": "武勋",
    "group_name": "分组",
    "join_time": "加入时间",
}


def auto_fit_column_width(ws):
    """自动调整列宽"""
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in col_cells:
            if cell.value:
                value_str = str(cell.value)
                length = 0
                for char in value_str:
                    if '\u4e00' <= char <= '\u9fff' or '\u3000' <= char <= '\u303f':
                        length += 2
                    else:
                        length += 1
                max_length = max(max_length, length)

        adjusted_width = min(max(max_length + 2, 8), 40)
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = adjusted_width


def apply_header_style(ws):
    """应用表头样式"""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border


def apply_data_style(ws):
    """应用数据区域样式"""
    data_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = data_alignment
            cell.border = thin_border


def _get_where_sql(filter_valid: bool, filter_no_npc: bool) -> str:
    """构建 WHERE 子句"""
    where_clauses = []
    if filter_valid:
        where_clauses.append("attack_hero1_id != 0 AND attack_hero2_id != 0 AND attack_hero3_id != 0")
    if filter_no_npc:
        where_clauses.append("npc != 1")
    return " WHERE " + " AND ".join(where_clauses) if where_clauses else ""


def _build_query(base_query: str, where_sql: str, additional_condition: str = None) -> str:
    """构建完整查询语句"""
    if additional_condition:
        if where_sql:
            return f"{base_query}{where_sql} AND {additional_condition}"
        else:
            return f"{base_query} WHERE {additional_condition}"
    return f"{base_query}{where_sql}"


def _create_worksheet(wb, title, field_names, rows):
    """创建工作表"""
    ws = wb.create_sheet(title) if wb.sheetnames and wb.sheetnames[0] != title else wb.active
    if ws.title != title:
        ws.title = title

    ws.append(field_names)
    for row in rows:
        ws.append(list(row))

    apply_header_style(ws)
    apply_data_style(ws)
    auto_fit_column_width(ws)
    ws.freeze_panes = "A2"

    return ws


def export_to_excel(db_path: str, output_path: str,
                    filter_valid: bool = False, filter_no_npc: bool = False,
                    include_members: bool = True) -> int:
    """
    导出数据到单个 Excel 文件
    返回导出的战报数量
    """
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    wb = openpyxl.Workbook()

    # 战报表
    fields = list(BATTLE_REPORT_FIELDS.keys())
    field_names = list(BATTLE_REPORT_FIELDS.values())
    where_sql = _get_where_sql(filter_valid, filter_no_npc)

    cursor.execute(f"SELECT {', '.join(fields)} FROM stzb_battle_reports{where_sql} ORDER BY battle_time DESC")
    rows = cursor.fetchall()

    _create_worksheet(wb, "战报数据", field_names, rows)
    report_count = len(rows)

    # 成员表
    if include_members:
        member_fields = list(TEAM_MEMBER_FIELDS.keys())
        member_field_names = list(TEAM_MEMBER_FIELDS.values())
        cursor.execute(f"SELECT {', '.join(member_fields)} FROM stzb_team_members ORDER BY power DESC")
        members = cursor.fetchall()
        _create_worksheet(wb, "同盟成员", member_field_names, members)

    wb.save(output_path)
    conn.close()

    return report_count


def export_to_excel_by_alliance(db_path: str, output_dir: str,
                                filter_valid: bool = False, filter_no_npc: bool = False,
                                include_members: bool = True) -> dict:
    """
    按同盟分组导出到独立的 Excel 文件
    返回: {alliance_name: file_path}
    """
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    where_sql = _get_where_sql(filter_valid, filter_no_npc)

    # 获取所有同盟列表
    base_query = "SELECT DISTINCT attack_union_name FROM stzb_battle_reports"
    cursor.execute(_build_query(base_query, where_sql))
    attack_alliances = [row[0] for row in cursor.fetchall() if row[0] and row[0].strip()]

    base_query = "SELECT DISTINCT defend_union_name FROM stzb_battle_reports"
    cursor.execute(_build_query(base_query, where_sql))
    defend_alliances = [row[0] for row in cursor.fetchall() if row[0] and row[0].strip()]

    alliances = list(set(attack_alliances + defend_alliances))

    # 准备字段
    fields = list(BATTLE_REPORT_FIELDS.keys())
    field_names = list(BATTLE_REPORT_FIELDS.values())
    member_fields = list(TEAM_MEMBER_FIELDS.keys())
    member_field_names = list(TEAM_MEMBER_FIELDS.values())

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    exported_files = {}

    for alliance in alliances:
        # 查询该同盟的所有战报
        alliance_condition = "(attack_union_name = ? OR defend_union_name = ?)"
        base_query = f"SELECT {', '.join(fields)} FROM stzb_battle_reports"
        query = _build_query(base_query, where_sql, alliance_condition) + " ORDER BY battle_time DESC"
        cursor.execute(query, (alliance, alliance))
        rows = cursor.fetchall()

        if not rows:
            continue

        # 创建工作簿
        wb = openpyxl.Workbook()
        _create_worksheet(wb, "战报数据", field_names, rows)

        # 成员表
        if include_members:
            cursor.execute(
                f"SELECT {', '.join(member_fields)} FROM stzb_team_members WHERE group_name = ? ORDER BY power DESC",
                (alliance,)
            )
            members = cursor.fetchall()
            _create_worksheet(wb, "同盟成员", member_field_names, members)

        # 保存文件（文件名中去掉特殊字符）
        safe_name = alliance.replace("/", "_").replace("\\", "_").replace(":", "_").replace("*", "_").replace("?", "_")
        file_path = output_path / f"{safe_name}.xlsx"
        wb.save(str(file_path))

        exported_files[alliance] = str(file_path)

    conn.close()

    return exported_files
